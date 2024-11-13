import openpyxl

from datetime import datetime

# 获取今天的日期
today = datetime.now()

# 格式化日期为 YYYYMMDD 格式
formatted_date = today.strftime('%Y%m%d')

# 打开源Excel文件
source_file_path = '/Users/enmoe/Downloads/xxxx.xlsx'
source_workbook = openpyxl.load_workbook(source_file_path)
source_sheet = source_workbook.active

# 打开目标Excel文件
target_file_path = '/Users/enmoe/Documents/开通台账/开通台账模版.xlsx'
target_workbook = openpyxl.load_workbook(target_file_path)
target_sheet = target_workbook.active

# 获取指定单元格的值
apply_unit = source_sheet['C2'].value
implement_contact = source_sheet['N3'].value
contact_phone = source_sheet['Q3'].value
project_name = source_sheet['E16'].value
apply_unit_connect_person = source_sheet['C3'].value
apply_unit_connect_info = source_sheet['G3'].value

system_name = ''

name = 'XXX'
ssh_port = '22'

zone = ''
num = 0
start_row = 3  # 从第三行开始添加数据

# 从第7行开始遍历每一行，直到最后一行
for row in source_sheet.iter_rows(min_row=7, max_row=source_sheet.max_row):
    # 检查这一行是否为空
    if any(cell.value for cell in row):
        # 打印这一行的值
        if '注:安全服务选项' not in row[0].value:
            num += 1
            if row[0].value == '政务外网':
                zone = 'ZWW'
            elif row[0].value == '互联网':
                zone = 'HLW'
            
            ip_addr = row[1].value
            vm_name = ip_addr.replace('.', '-')
            if apply_unit == 'XXX':
                vm_name = 'XXX-' + vm_name
            elif apply_unit == 'XXXX':
                vm_name = 'XXXX-' + vm_name
                
            cpu_type = row[5].value
            cpu_core = row[6].value
            memory_size = row[7].value
            system_disk = row[11].value
            other_disk = row[13].value
            os_version = row[9].value
            sec_type = row[8].value
            system_name = row[2].value

            info = [
                str(num),
                '是',
                vm_name,
                '',
                zone,
                ip_addr,
                ssh_port,
                'root',
                '',
                cpu_type,
                str(cpu_core),
                str(memory_size),
                str(system_disk),
                str(other_disk),
                os_version,
                sec_type,
                '',
                '',
                name,
                formatted_date,
                '/',
                system_name,
                apply_unit,
                apply_unit_connect_person,
                apply_unit_connect_info,
                implement_contact,
                contact_phone
                
                
                
            ]

            # 将info内容添加到目标Excel文件的对应行
            for col_idx, value in enumerate(info, start=1):
                target_sheet.cell(row=start_row, column=col_idx, value=value)

            start_row += 1

file_name = apply_unit + '-' + system_name + '-XCY-' + formatted_date + '-JF.xlsx'

out_file_path = '/Users/enmoe/Documents/开通台账/' + file_name

# 保存目标Excel文件
target_workbook.save(out_file_path)
